__author__ = 'Jens'

"""
    Requires:
        + Python 2.7
        + pyWinUSB
        + pyOCD (mbedmicro/pyOCD on github)
"""

from time import sleep
import random
import os
from pyOCD.board import MbedBoard
import logging
from openpyxl import load_workbook
from instruction.instruction import Instruction

# Class to encapsulate all behavior to inject faults into registers
class RegisterFaultInjector:
    def __init__(self, board=None):
        if board is None:
             self.board = self._initialize()
             # get the target
             self.target = self.board.target
        else:
            self.board = board
            self.target = self.board.target
        # initialize excel workbook
        self.workBook = load_workbook('{}/{}'.format(os.path.dirname(os.path.dirname(__file__)), 'resultsFaultInjection.xlsx'))
        # Variables to detect type of fault
        self.hardFault = 0
        self.detectedFault = 0


    # Public methods
    def unInit(self):
        if self.board != None:
            logging.info("Board uninitialised")
            self.board.uninit()

    def injectionSuite(self, tries, nameSheet):
        """
            Function that injects fault for every possible bit to flip.
            Tries: number of injections per bit-flip location
        """
        # Initialize workSheet
        ws = self._initExcelSheet(nameSheet, tries)
        # Initialize some variables
        faultsInjected = 0
        suiteEffects = 0
        suiteNoEffects = 0
        bitToFlip = 0x00000001
        rowVar = 2
        # Loop over all possible bits that can be flipped
        for i in range(0,32):
            loopMessage = "Injection round for bitTofFlip: 0x%X" % bitToFlip
            logging.log(25, loopMessage)
            #variable for row in excelSheet
            rowVar += 1
            effects = 0
            noEffects = 0
            self.hardFault = 0
            self.detectedFault = 0
            for j in range(0,tries):
                logging.log(25, "New try: %s" % j)
                success = self.injectFaultRegisterInstructionAware(bitToFlip)
                if success:
                    faultsInjected += 1
                    # sleep for small amount, needed so watchdogTimer can run out
                    sleep(0.3)
                    # Check for error
                    error = self.checkError()
                    if error == 1:
                        suiteEffects += 1
                        effects += 1
                    else:
                        suiteNoEffects +=1
                        noEffects += 1
                    sleep(0.1)
                # Reset device, just to be sure
                self.resetDevice()
                # Sleep random time, to inject next fault into random instruction
                sleepTime = 0.1 + random.random() * 0.5
                sleep(sleepTime)
            # write results to Excel file
            ws = self._writeToExcel(self.workBook, ws, nameSheet, rowVar, ("0x%X" % bitToFlip), effects, noEffects)
            # Move bit to flip 1 position to the left
            bitToFlip = bitToFlip << 1
        rowVar += 1
        ws = self._writeToExcel(self.workBook, ws, nameSheet, rowVar, faultsInjected, suiteEffects, suiteNoEffects)
        self._logFinalResults(faultsInjected, suiteEffects, suiteNoEffects)

    def checkError(self):
        """
            Function to check if injected fault generated an error
            Checks if all 4 leds are on
        """
        error = 0
        try:
            # Halt mbed
            self.target.halt()
            # Read register of leds
            status = self.target.readMemory(0x2009C038)
            # Resume execution
            self.target.resume()
            # Isolate bits of Leds
            Led3 = (status >> 21) & 1
            Led4 = (status >> 23) & 1
            # Generate evaluation
            error = Led3 | Led4
            if (error == 1):
                logging.info("------------- ERROR --------------")
                if Led4 == 1:
                    self.hardFault +=1
                elif Led3 == 1:
                    self.detectedFault += 1
        except Exception as e:
            print("! -- ! Exception: %s" %e)
        finally:
            return error

    def injectFaultRegisterInstructionAware(self, bitToFlip = 0x00000008, flagAddress=None):
        """
        Function to inject a fault in a register that is chosen
        by analysing the current instruction
        :return: True if fault was injected, False if not.
        """
        try:
            # Halt mbed execution
            self.target.halt()
            # Get the instruction
            if self._checkFlag(flagAddress):
                instruction = Instruction(self.target)
                reg = instruction.getNecessaryRegister()
                if reg is not None:
                    self._injectFaultRegister(reg,bitToFlip)
                    # Resume mbed execution
                    success = True
                    logging.log(25, "Successfully injected a fault into %s" % reg)
                else:
                    success = False
            else:
                logging.log(25, "Flag at address 0x%X was 0, No fault injected!" % flagAddress)
                success = False
            self.target.resume()
            return success
        except Exception as e:
            logging.error("Could not inject a fault into {}!\n{}".format(reg, e))

    def injectFaultRegisters(self, listOfRegs, bitToFlip = 0x00001000):
        """
            listOfRegs: The list of registers whose content must be bit-flipped. Eg: ['r0', 'r2']
            bitToFlip: Mask to be used to flip bit. Default: 0x00001000.
        """
        try:
            logging.info ("Starting loop to inject bit-flip. Ctrl-C or reset to end.")
            while True:
                # Loop to keep injecting bit flip into the registers of 'listOfRegs'
                for i in range(len(listOfRegs)):
                    # Halt target
                    self.target.halt()
                    # Inject fault in the currently selected register
                    self._injectFaultRegister(listOfRegs[i], bitToFlip)
                    # Resume execution
                    self.target.resume()
                    # Wait to inject next fault
                    sleep(0.1)
        except Exception as e:
            print("! -- ! Exception: %s" %e)

    def resetDevice(self):
        """
            Function to reset the device
        """
        try:
            self.target.halt()
            logging.info("Resetting device")
            self.target.reset()
            self.target.resume()
        except Exception as e:
            print("! -- ! Exception: %s" %e)

    def setBASEPRI(self, value):
        self.target.halt()
        current = self.target.readCoreRegister('basepri')
        self.target.writeCoreRegister('basepri', value)
        self.target.resume()
        logging.log(25, "BASEPRI was 0x{:X}, changed to 0x{:X}".format(current, value))

    def setFAULTMASK(self, value):
        self.target.halt()
        current = self.target.readCoreRegister('faultmask')
        self.target.writeCoreRegister('faultmask', value)
        self.target.resume()
        logging.log(25, "FAULTMASK was 0x{:X}, changed to 0x{:X}".format(current,value))

    # Private methods
    def _initialize(self):
        """
            Used to initialize the mbed
        """
        board = None
        try:
            # Search mbed and initialize
            board = MbedBoard.chooseBoard()
            target_type = board.getTargetType()

            # Needed to write to memory
            if target_type == "lpc1768":
                addr = 0x10000001
                size = 0x1102
                addr_flash = 0x10000
            else:
                raise Exception("Only lpc_1768 supported for this small fault injector")

            # Get everything that is needed
            #target = board.target
            #transport = board.transport
            #interface = board.interface
            #flash = board.flash

        finally:
            if board != None:
                return board

    def _injectFaultRegister(self, reg, bitToFlip=0x00000008):
        """
            Function to inject a fault in the given register (reg)
            on the position defined by bitToFlip

            Target must be Halted!!!!
        """
        logging.info("Start to inject fault in register %s" % reg)
        try:
            # Read and print content of register
            content = self.target.readCoreRegister(reg)
            logging.debug("Content of register: 0x%X" % content)
            # Bit-flip content by XOR with bitToFlip
            content ^= bitToFlip
            logging.debug("New content: 0x%X" % content)
            # Write new content to register
            self.target.writeCoreRegister(reg, content)
        except Exception as e:
            print("! -- ! Exception: %s" %e)

    def _checkFlag(self, flagAddress):
        if flagAddress is not None:
            try:
                flag = self.target.readMemory(flagAddress)
                flag &= 1
                if flag == 1:
                    return True
                else:
                    return False
            except Exception as e:
                logging.error("Could not read the flag at address 0x%X", flagAddress)
                return False
        else:
            return True

    def _initExcelSheet(self, nameSheet, tries):
        # Create a new worksheet
        ws = self.workBook.create_sheet()
        # Set title
        ws.title = nameSheet
        # A1 holds info about tries/bit
        ws['A1'] = ("Results of the fault injection, with %s tries per bit" % tries)
        # Fill in row 2, head of table
        ws.cell(row = 2, column = 1).value = "Bitmask"
        ws.cell(row = 2, column = 2).value = "Effect"
        ws.cell(row = 2, column = 3).value = "No effect"
        ws.cell(row = 2, column = 4).value = "Efficiency [%]"
        ws.cell(row = 2, column = 5).value = "Hardfaults"
        ws.cell(row = 2, column = 6).value = "Detectedfaults"
        # Return ws
        return ws

    def _writeToExcel(self, workBook, workSheet, nameWorkSheet, rowVar, bitToFlip, effects, noEffects):
        # Fill content into cells
        workSheet.cell(row = rowVar, column = 1).value = bitToFlip
        workSheet.cell(row = rowVar, column = 2).value = effects
        workSheet.cell(row = rowVar, column = 3).value = noEffects
        formula = "=(B{}/(B{} + C{}))*100".format(rowVar, rowVar, rowVar)
        workSheet.cell(row = rowVar, column = 4).value = (formula)
        workSheet.cell(row = rowVar, column = 5).value = self.hardFault
        workSheet.cell(row = rowVar, column = 6).value = self.detectedFault
        # Save workbook
        workBook.save('resultsFaultInjection.xlsx')
        # return worksheet
        return workBook.get_sheet_by_name(nameWorkSheet)

    def _logFinalResults(self, faultsInjected, suiteEffects, suiteNoEffects):
        logging.log(25, "-----------------------------------------------")
        logging.log(25, "Number of faults injected: %s" % faultsInjected)
        logging.log(25, "Faults that had an effect: %s" % suiteEffects)
        logging.log(25, "Faults that had no effect: %s" % suiteNoEffects)
        logging.log(25, "Efficiency: %f" % (float(suiteEffects)/float(faultsInjected)))
        logging.log(25, "-----------------------------------------------")
